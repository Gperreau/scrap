"""
Scraper ORIAS — Base Complète (27 000 lignes)
=============================================
Pipeline ultra-optimisé en 3 phases :
  Phase 1 — SIRENE batch   : SIREN → nom + URL (async, 50 conc., ~2 min)
  Phase 2 — Scraping async : HTTP aiohttp avec pool de 20 conns simultanées
  Phase 3 — DDG ciblé      : uniquement pour les entrées sans URL (~70%)

SOUS-COMMANDES
  run    : traitement d'un bloc (une instance ou un job GitHub Actions)
  split  : découpe le CSV + génère le workflow GitHub Actions matrix
  merge  : fusionne N fichiers Excel de résultats
  stats  : affiche les statistiques du cache sans relancer

PRÉREQUIS
  pip install aiohttp aiofiles requests beautifulsoup4 lxml openpyxl
              duckduckgo-search tqdm pandas

FORMAT D'ENTRÉE (CSV ou Excel — export ORIAS)
  Colonnes attendues (auto-détectées) :
    numéro orias  |  siren  |  [raison sociale]  |  [état]

USAGE RAPIDE
  # 1. Découper + générer le workflow GitHub Actions
  python orias_scraper.py split --input orias_coa.csv --chunks 9

  # 2. Pousser sur GitHub et déclencher le workflow (ou lancer localement)
  python orias_scraper.py run --input orias_coa.csv

  # 3. Fusionner les résultats après exécution parallèle
  python orias_scraper.py merge --pattern "results_chunk_*.xlsx"
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import logging
import os
import re
import smtplib
import ssl
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from email.message import EmailMessage
from email.utils import formatdate
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import aiohttp
    HAS_AIOHTTP = True
except ImportError:
    HAS_AIOHTTP = False
    print("⚠  pip install aiohttp aiofiles")

try:
    from ddgs import DDGS
    HAS_DDG = True
except ImportError:
    HAS_DDG = False

try:
    from tqdm import tqdm
    from tqdm.asyncio import tqdm as atqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

logging.basicConfig(level=logging.WARNING)

# ═══════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════

SIRENE_API   = "https://recherche-entreprises.api.gouv.fr/search"
ANNUAIRE_URL = "https://annuaire-entreprises.data.gouv.fr/etablissement/{siret}"
HTTP_TIMEOUT = 14          # secondes
HTTP_CONCURRENT = 20       # connexions HTTP simultanées
SIRENE_CONCURRENT = 50     # appels SIRENE simultanées (pas de rate-limit)
DDG_DELAY = 1.5            # secondes entre requêtes DDG
SAVE_EVERY = 100           # entrées entre chaque sauvegarde du cache

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

EXCLUDED_DOMAINS = [
    "linkedin.com", "societe.com", "verif.com", "infogreffe.fr",
    "pappers.fr", "kompass.com", "facebook.com", "twitter.com",
    "youtube.com", "pagesjaunes.fr", "mappy.com", "annuaire-entreprises.data.gouv.fr",
    "score3.fr", "hoodspot.fr", "manageo.fr", "corporama.com",
]

# ═══════════════════════════════════════════════════════════
#  DÉTECTION PRODUITS — RÈGLES ÉLARGIES
# ═══════════════════════════════════════════════════════════

SANTE_BASE = [
    "santé", "sante", "mutuelle", "complémentaire santé", "assurance santé",
    "frais de santé", "remboursement soins", "hospitalisation",
    "optique dentaire", "médecines douces", "couverture médicale",
    "garantie santé", "surcomplémentaire",
]
SANTE_HINTS = [
    "particulier", "individuel", "individuelle",
    "tns", "travailleur non salarié", "non salarié",
    "entrepreneur", "auto-entrepreneur", "indépendant",
    "libéral", "profession libérale", "artisan", "commerçant",
    "gérant", "dirigeant", "chef d'entreprise",
    "retraité", "senior", "étudiant", "famille", "conjoint",
    "pour vous", "votre santé", "votre mutuelle",
    "devis mutuelle", "loi madelin",
]
PERIN_KW = [
    "retraite",
    " per ", " per,", " per.", " per)", "(per ", "per\n",
    "perin", "perp",
    "plan épargne retraite", "plan epargne retraite",
    "épargne retraite", "epargne retraite",
    "madelin", "loi madelin",
    "préparer sa retraite", "préparer votre retraite",
    "preparer sa retraite", "preparer votre retraite",
    "retraite supplémentaire", "retraite complementaire",
    "sortie en rente", "sortie en capital",
    "placement retraite", "patrimoine",
]


def detect_sante(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in SANTE_BASE) and any(h in t for h in SANTE_HINTS)


def detect_perin(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in PERIN_KW)


# ═══════════════════════════════════════════════════════════
#  PHASE 1 — BATCH SIRENE (async)
# ═══════════════════════════════════════════════════════════

async def sirene_one(session: aiohttp.ClientSession, siren: str, sem: asyncio.Semaphore) -> dict:
    """Interroge l'API SIRENE pour un SIREN. Retourne {} si non trouvé."""
    s = re.sub(r"[\s\-\.]", "", siren)
    if len(s) not in (9, 14):
        return {}
    # Utilise les 9 premiers chiffres (SIREN)
    s9 = s[:9]
    url = f"{SIRENE_API}?q={s9}&mtd=true"
    try:
        async with sem:
            async with session.get(url, headers=HEADERS, timeout=aiohttp.ClientTimeout(total=10)) as r:
                if r.status != 200:
                    return {}
                data = await r.json()
                hits = data.get("results", [])
                if not hits:
                    return {}
                h     = hits[0]
                siege = h.get("siege", {})
                return {
                    "siren":     s9,
                    "nom":       h.get("nom_complet", ""),
                    "url":       siege.get("site_internet", "").strip(),
                    "ape":       h.get("activite_principale", ""),
                    "ape_label": h.get("libelle_activite_principale", ""),
                    "ville":     siege.get("libelle_commune", ""),
                    "cp":        siege.get("code_postal", ""),
                }
    except Exception:
        return {}


async def sirene_batch(rows: list[dict]) -> dict[str, dict]:
    """
    Interroge SIRENE pour tous les SIRENs en parallèle.
    Retourne {siren: sirene_data}.
    ~2 min pour 27 000 entrées (50 connexions simultanées, sans rate-limit).
    """
    if not HAS_AIOHTTP:
        # Fallback synchrone
        results = {}
        for row in rows:
            data = _sirene_sync(row.get("siren", ""))
            if data:
                results[data["siren"]] = data
        return results

    sem     = asyncio.Semaphore(SIRENE_CONCURRENT)
    results = {}
    conn    = aiohttp.TCPConnector(limit=SIRENE_CONCURRENT, ttl_dns_cache=300)

    async with aiohttp.ClientSession(connector=conn) as session:
        tasks = {
            row.get("siren", ""): sirene_one(session, row.get("siren", ""), sem)
            for row in rows if row.get("siren")
        }
        bar = None
        if HAS_TQDM:
            bar = tqdm(total=len(tasks), desc="SIRENE batch", unit="siren", ncols=80)

        for siren, coro in tasks.items():
            data = await coro
            if data:
                results[data["siren"]] = data
            if bar:
                bar.update(1)

        if bar:
            bar.close()

    nb_url = sum(1 for d in results.values() if d.get("url"))
    print(f"  → {len(results)} réponses SIRENE, {nb_url} avec URL directe.")
    return results


def _sirene_sync(siren: str) -> dict:
    """Version synchrone de sirene_one (fallback sans aiohttp)."""
    s = re.sub(r"[\s\-\.]", "", siren)[:9]
    try:
        r = requests.get(f"{SIRENE_API}?q={s}&mtd=true", headers=HEADERS, timeout=10)
        if r.status_code != 200:
            return {}
        hits = r.json().get("results", [])
        if not hits:
            return {}
        h     = hits[0]
        siege = h.get("siege", {})
        return {
            "siren": s, "nom": h.get("nom_complet", ""),
            "url": siege.get("site_internet", "").strip(),
            "ape": h.get("activite_principale", ""),
            "ape_label": h.get("libelle_activite_principale", ""),
            "ville": siege.get("libelle_commune", ""),
        }
    except Exception:
        return {}


# ═══════════════════════════════════════════════════════════
#  PHASE 2 — SCRAPING ASYNC
# ═══════════════════════════════════════════════════════════

async def scrape_url_async(session: aiohttp.ClientSession, url: str, sem: asyncio.Semaphore) -> str:
    """Scrappe une URL et retourne le texte visible (lowercase)."""
    try:
        async with sem:
            async with session.get(
                url, headers=HEADERS,
                timeout=aiohttp.ClientTimeout(total=HTTP_TIMEOUT),
                allow_redirects=True,
            ) as r:
                if r.status != 200:
                    return ""
                html = await r.text(errors="replace")
        soup = BeautifulSoup(html, "lxml")
        for tag in soup(["script", "style", "nav", "footer", "head", "noscript"]):
            tag.decompose()
        return soup.get_text(separator=" ", strip=True).lower()
    except Exception:
        return ""


async def scrape_batch(urls: dict[str, str]) -> dict[str, str]:
    """
    Scrappe un dict {id: url} en parallèle.
    Retourne {id: texte}.
    """
    if not HAS_AIOHTTP or not urls:
        return {}
    sem  = asyncio.Semaphore(HTTP_CONCURRENT)
    conn = aiohttp.TCPConnector(limit=HTTP_CONCURRENT, ttl_dns_cache=300)
    results = {}

    async with aiohttp.ClientSession(connector=conn) as session:
        tasks = {uid: scrape_url_async(session, url, sem) for uid, url in urls.items()}
        bar = tqdm(total=len(tasks), desc="Scraping URLs", unit="page", ncols=80) if HAS_TQDM else None

        for uid, coro in tasks.items():
            text = await coro
            results[uid] = text
            if bar:
                bar.update(1)

        if bar:
            bar.close()

    return results


# ═══════════════════════════════════════════════════════════
#  PHASE 3 — DDG (synchrone avec verrou)
# ═══════════════════════════════════════════════════════════

_ddg_lock      = threading.Lock()
_last_ddg_call = 0.0


def ddg_search(query: str, max_results: int = 8, delay: float = DDG_DELAY) -> list[dict]:
    global _last_ddg_call
    if not HAS_DDG:
        return []
    with _ddg_lock:
        wait = delay - (time.time() - _last_ddg_call)
        if wait > 0:
            time.sleep(wait)
        try:
            with DDGS() as d:
                results = list(d.text(query, max_results=max_results))
            _last_ddg_call = time.time()
            return results
        except Exception:
            _last_ddg_call = time.time()
            return []


def find_url_ddg(nom: str, siren: str, delay: float = DDG_DELAY) -> tuple[str, str]:
    """
    Cherche l'URL officielle via DDG.
    Essaie dans l'ordre :
      1. SIREN brut (souvent très précis)
      2. Nom entre guillemets + "courtier assurance"
    Retourne (url, méthode).
    """
    queries = []
    if siren:
        queries.append((f'"{siren}" assurance courtier', "DDG SIREN"))
    if nom:
        queries.append((f'"{nom}" courtier assurance', "DDG nom"))
        queries.append((f'{nom} assurance site officiel', "DDG générique"))

    for query, method in queries:
        results = ddg_search(query, max_results=6, delay=delay)
        for r in results:
            url = r.get("href", "")
            if url and not any(e in url for e in EXCLUDED_DOMAINS):
                return url, method
    return "", "Non trouvé"


def scrape_url_sync(url: str) -> str:
    """Version synchrone pour le fallback sans aiohttp."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=HTTP_TIMEOUT, allow_redirects=True)
        if r.status_code != 200:
            return ""
        soup = BeautifulSoup(r.content, "lxml")
        for tag in soup(["script", "style", "nav", "footer", "head", "noscript"]):
            tag.decompose()
        return soup.get_text(separator=" ", strip=True).lower()
    except Exception:
        return ""


# ═══════════════════════════════════════════════════════════
#  ANALYSE D'UNE ENTRÉE
# ═══════════════════════════════════════════════════════════

def build_result(
    orias: str, nom: str, siren: str,
    page_text: str, url: str, method: str,
    sirene_data: dict,
) -> dict:
    """Construit le dict résultat à partir des textes collectés."""
    combined = (page_text + " " + sirene_data.get("nom", "") + " " +
                sirene_data.get("ape_label", "")).lower()

    sante = detect_sante(combined)
    perin  = detect_perin(combined)

    contrib = []
    if page_text:
        if detect_sante(page_text):
            contrib.append(f"Page web ({method})")
        if detect_perin(page_text):
            if not contrib or "PERIN" not in str(contrib):
                contrib.append(f"Page web PERIN ({method})")
    if sirene_data.get("ape_label"):
        ape_text = sirene_data["ape_label"].lower()
        if detect_sante(ape_text):
            contrib.append("API SIRENE (Santé)")
        if detect_perin(ape_text):
            contrib.append("API SIRENE (PERIN)")

    score, detail = 0, []
    if url:
        score += 10; detail.append("URL trouvée (+10)")
    if page_text and len(page_text) > 1000:
        score += 30; detail.append("Page riche (+30)")
    elif page_text:
        score += 15; detail.append("Page courte (+15)")
    if sirene_data.get("nom"):
        score += 10; detail.append("SIRENE (+10)")
    if sante or perin:
        score += 20; detail.append("Produit détecté (+20)")
    if "DDG générique" in method:
        score = max(score - 5, 0)

    return {
        "numero_orias":      orias,
        "nom":               nom or sirene_data.get("nom", ""),
        "siren":             siren,
        "ape":               sirene_data.get("ape", ""),
        "ville":             sirene_data.get("ville", ""),
        "sante_individuelle": sante,
        "perin":             perin,
        "score":             min(score, 100),
        "niveau":            "Élevé ✅" if score >= 70 else "Moyen ⚠️" if score >= 40 else "Faible ❌",
        "url":               url,
        "methode":           method,
        "sources":           " | ".join(contrib) if contrib else "—",
        "detail_score":      " | ".join(detail),
    }


# ═══════════════════════════════════════════════════════════
#  ORCHESTRATION PRINCIPALE
# ═══════════════════════════════════════════════════════════

async def process_chunk(
    rows: list[dict],
    cache: dict,
    cache_path: str,
    delay: float,
    workers: int,
) -> list[dict]:
    """
    Traite un lot de lignes avec le pipeline en 3 phases.
    """
    # Séparer les entrées déjà en cache
    to_do   = [r for r in rows if r["orias"] not in cache]
    results = [cache[r["orias"]] for r in rows if r["orias"] in cache]

    print(f"  {len(results)} en cache | {len(to_do)} à traiter")
    if not to_do:
        return results

    # ── Phase 1 : SIRENE batch ───────────────────────────
    print(f"\n  Phase 1 — SIRENE batch ({len(to_do)} entrées)…")
    sirene_cache = await sirene_batch(to_do)

    # Répartition : avec URL directe vs sans
    with_url    = [(r, sirene_cache.get(re.sub(r'[\s\-\.]','',r['siren'])[:9], {}))
                   for r in to_do
                   if sirene_cache.get(re.sub(r'[\s\-\.]','',r['siren'])[:9], {}).get("url")]
    without_url = [(r, sirene_cache.get(re.sub(r'[\s\-\.]','',r['siren'])[:9], {}))
                   for r in to_do
                   if not sirene_cache.get(re.sub(r'[\s\-\.]','',r['siren'])[:9], {}).get("url")]

    nb_direct = len(with_url)
    nb_ddg    = len(without_url)
    est_ddg_h = nb_ddg * 2 * delay / 3600
    print(f"  → {nb_direct} URL directes (SIRENE) | {nb_ddg} via DDG (~{est_ddg_h:.1f}h)")

    # ── Phase 2 : Scraping async des URLs directes ───────
    page_texts: dict[str, str] = {}
    if with_url and HAS_AIOHTTP:
        print(f"\n  Phase 2 — Scraping async ({nb_direct} pages)…")
        url_map = {r["orias"]: sd["url"] for r, sd in with_url}
        page_texts = await scrape_batch(url_map)

    # Construire résultats pour les entrées avec URL directe
    for r, sd in with_url:
        text   = page_texts.get(r["orias"], "")
        result = build_result(r["orias"], r.get("nom",""), r["siren"],
                              text, sd["url"], "SIRENE direct", sd)
        cache[r["orias"]] = result
        results.append(result)

    # ── Phase 3 : DDG pour les entrées sans URL ───────────
    if without_url:
        print(f"\n  Phase 3 — DDG ({nb_ddg} entrées)…")
        counter = [0]
        bar     = tqdm(total=nb_ddg, desc="DDG search", unit="entrée", ncols=80) if HAS_TQDM else None

        def ddg_task(item):
            r, sd = item
            nom   = sd.get("nom") or r.get("nom", "")
            url, method = find_url_ddg(nom, r["siren"], delay)

            # Scraper si URL trouvée
            text = ""
            if url:
                text = scrape_url_sync(url)
                if len(text) < 300:
                    # Essayer sous-pages
                    for suf in ["/nos-offres", "/produits", "/services"]:
                        t2 = scrape_url_sync(url.rstrip("/") + suf)
                        if len(t2) > len(text):
                            text = t2

            result = build_result(r["orias"], nom, r["siren"], text, url, method, sd)
            counter[0] += 1
            if bar:
                bar.update(1)
            elif counter[0] % 50 == 0:
                print(f"    [{counter[0]}/{nb_ddg}]")

            return r["orias"], result

        # Exécuter DDG dans un ThreadPoolExecutor (I/O-bound, limité par rate limit)
        loop   = asyncio.get_event_loop()
        chunks = [without_url[i:i+100] for i in range(0, len(without_url), 100)]

        for chunk in chunks:
            with ThreadPoolExecutor(max_workers=min(workers, 3)) as executor:
                tasks = [loop.run_in_executor(executor, ddg_task, item) for item in chunk]
                done  = await asyncio.gather(*tasks, return_exceptions=True)

            for res in done:
                if isinstance(res, tuple):
                    orias_id, result = res
                    cache[orias_id] = result
                    results.append(result)

            # Sauvegarde intermédiaire après chaque tranche de 100
            _save_cache(cache, cache_path)

        if bar:
            bar.close()

    _save_cache(cache, cache_path)
    return results


# ═══════════════════════════════════════════════════════════
#  CACHE
# ═══════════════════════════════════════════════════════════

def _save_cache(cache: dict, path: str):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)
    os.replace(tmp, path)   # écriture atomique


def load_cache(path: str) -> dict:
    if Path(path).exists():
        try:
            return json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


# ═══════════════════════════════════════════════════════════
#  CHARGEMENT DU FICHIER SOURCE
# ═══════════════════════════════════════════════════════════

# Noms de colonnes courants dans l'export ORIAS
ORIAS_COL_CANDIDATES = ["numero orias", "numéro orias", "n° orias", "orias", "num_orias",
                         "numero_orias", "n_orias", "immatriculation"]
SIREN_COL_CANDIDATES = ["siren", "siret", "numero siren", "numéro siren",
                         "n° siren", "siren/siret", "num_siren"]
NOM_COL_CANDIDATES   = ["raison sociale", "raison_sociale", "nom", "denomination",
                         "nom courtier", "denominationsociale"]


def _find_col(col_map: dict, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in col_map:
            return col_map[c]
    return None


def load_input(
    path: str,
    col_orias: str = "",
    col_siren: str = "",
    col_nom: str = "",
    chunk_id: int = 0,
    total_chunks: int = 1,
) -> list[dict]:
    """
    Charge le fichier et retourne uniquement le bloc (chunk) de cette instance.
    chunk_id  : 1-based (1, 2, … total_chunks)
    total_chunks : nombre total de blocs (pour GitHub Actions matrix)
    """
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xls", ".xlsm"):
        if not HAS_PANDAS:
            raise RuntimeError("pip install pandas openpyxl")
        df = pd.read_excel(path, dtype=str)
        df.columns = df.columns.str.strip().str.lower()
        raw = df.to_dict("records")
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            # Auto-détection du séparateur
            sample = f.read(4096); f.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            reader = csv.DictReader(f, dialect=dialect)
            raw = [
                {k.strip().lower(): (v or "").strip() for k, v in row.items()}
                for row in reader
            ]

    if not raw:
        raise ValueError("Fichier vide.")

    col_map = {k.lower(): k for k in raw[0].keys()}

    real_orias = _find_col(col_map, [col_orias.lower()] + ORIAS_COL_CANDIDATES if col_orias else ORIAS_COL_CANDIDATES)
    real_siren = _find_col(col_map, [col_siren.lower()] + SIREN_COL_CANDIDATES if col_siren else SIREN_COL_CANDIDATES)
    real_nom   = _find_col(col_map, [col_nom.lower()]   + NOM_COL_CANDIDATES   if col_nom   else NOM_COL_CANDIDATES)

    if not real_orias and not real_siren:
        raise ValueError(
            f"Colonnes ORIAS/SIREN introuvables. Colonnes disponibles : {list(raw[0].keys())}\n"
            "Utilisez --col-orias et --col-siren pour les préciser."
        )

    rows = []
    for r in raw:
        # Conversion sécurisée : float/None/nan → string propre
        def _str(val) -> str:
            if val is None:
                return ""
            s = str(val).strip()
            # pandas lit les cellules vides comme "nan" et les entiers comme "362521879.0"
            if s.lower() in ("nan", "none", ""):
                return ""
            # Supprimer le .0 final des nombres lus comme float
            if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
                s = s[:-2]
            return s

        orias = _str(r.get(real_orias)) if real_orias else ""
        siren = _str(r.get(real_siren)) if real_siren else ""
        nom   = _str(r.get(real_nom))   if real_nom   else ""
        siren = re.sub(r"[\s\-\.]", "", siren)

        if not orias and not siren:
            continue

        # Filtre état si présent (ne garder que les immatriculations valides)
        etat_raw = r.get("état", r.get("etat", r.get("statut", "")))
        etat = _str(etat_raw).lower()
        if etat and "supprim" in etat:
            continue

        rows.append({"orias": orias or siren, "siren": siren, "nom": nom})

    total = len(rows)
    print(f"  → {total} entrées valides chargées")

    # Découpage pour le mode matrix
    if total_chunks > 1 and chunk_id > 0:
        size  = -(-total // total_chunks)   # division par excès
        start = (chunk_id - 1) * size
        end   = min(start + size, total)
        rows  = rows[start:end]
        print(f"  → Bloc {chunk_id}/{total_chunks} : lignes {start+1}–{end} ({len(rows)} entrées)")

    return rows


# ═══════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ═══════════════════════════════════════════════════════════

def build_excel(results: list[dict], path: str, chunk_id: int = 0, total_chunks: int = 1):
    C_BG  = "1F3864"; C_ACC = "2E6DA4"
    C_ODD = "EEF2F7"; C_EV  = "FFFFFF"
    C_HI  = "1A9641"; C_MD  = "E8A000"; C_LO  = "C0392B"
    C_OUI = "D5F5E3"; C_NON = "FADBD8"
    C_OUF = "1A7A45"; C_NOF = "922B21"

    def brd():
        s = Side(style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    def scol(s):
        return C_HI if s >= 70 else C_MD if s >= 40 else C_LO

    n    = len(results)
    nb_s = sum(1 for r in results if r.get("sante_individuelle"))
    nb_p = sum(1 for r in results if r.get("perin"))
    nb_u = sum(1 for r in results if r.get("url"))

    wb = Workbook(); ws = wb.active; ws.title = "Résultats ORIAS"
    NC = 10

    chunk_label = f" — Bloc {chunk_id}/{total_chunks}" if total_chunks > 1 else ""

    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    c = ws["A1"]
    c.value = f"ORIAS — Santé Individuelle & PERIN{chunk_label}"
    c.font  = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=C_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(NC)}2")
    c = ws["A2"]
    c.value = (f"Généré le {time.strftime('%d/%m/%Y %H:%M')}  —  "
               f"{n} courtiers  |  {nb_u} pages trouvées  |  "
               f"Santé : {nb_s}  |  PERIN : {nb_p}  —  Règles élargies")
    c.font  = Font(name="Arial", size=8, italic=True, color="555555")
    c.fill  = PatternFill("solid", fgColor="D9E2F0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = [
        "N° ORIAS", "SIREN", "Raison Sociale", "Ville / APE",
        "Santé Individuelle", "PERIN",
        "Score", "Niveau", "URL", "Sources / Méthode",
    ]
    widths = [14, 12, 32, 24, 18, 14, 10, 14, 50, 55]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=C_ACC)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 32

    sorted_res = sorted(
        results,
        key=lambda r: (-(int(r.get("sante_individuelle", False)) + int(r.get("perin", False))),
                       -r.get("score", 0))
    )

    for ri, e in enumerate(sorted_res, 4):
        bg    = C_ODD if ri % 2 == 0 else C_EV
        score = e.get("score", 0)
        sante = e.get("sante_individuelle", False)
        perin = e.get("perin", False)

        vals = [
            e.get("numero_orias", ""),
            e.get("siren", ""),
            e.get("nom", ""),
            f"{e.get('ville','')} — {e.get('ape','')}".strip(" —"),
            "✅ Oui" if sante else "❌ Non",
            "✅ Oui" if perin  else "❌ Non",
            score,
            e.get("niveau", ""),
            e.get("url", ""),
            f"{e.get('sources','')} | {e.get('methode','')}".strip(" |"),
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = brd()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if ci in (1, 2, 3, 4):
                c.font = Font(name="Arial", size=9, bold=(ci == 1)); c.fill = PatternFill("solid", fgColor=bg)
            elif ci == 5:
                c.font = Font(name="Arial", size=9, bold=True, color=C_OUF if sante else C_NOF)
                c.fill = PatternFill("solid", fgColor=C_OUI if sante else C_NON)
            elif ci == 6:
                c.font = Font(name="Arial", size=9, bold=True, color=C_OUF if perin else C_NOF)
                c.fill = PatternFill("solid", fgColor=C_OUI if perin else C_NON)
            elif ci == 7:
                c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=scol(score))
            elif ci == 8:
                c.font = Font(name="Arial", size=9, bold=True, color=scol(score)); c.fill = PatternFill("solid", fgColor=bg)
            elif ci == 9:
                c.font = Font(name="Arial", size=8, color="1155CC" if val else "999999")
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                c.font = Font(name="Arial", size=8, color="555555"); c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A4"
    wb.save(path)
    print(f"\n✅ Excel : {path}  ({nb_s} Santé | {nb_p} PERIN)")


# ═══════════════════════════════════════════════════════════
#  SPLIT + GITHUB ACTIONS WORKFLOW
# ═══════════════════════════════════════════════════════════

def split_and_generate_workflow(
    input_path: str,
    n_chunks: int,
    script_name: str = "orias_scraper.py",
    col_orias: str = "",
    col_siren: str = "",
    col_nom: str = "",
    delay: float = DDG_DELAY,
):
    """
    Divise le fichier source en N blocs CSV et génère le workflow
    GitHub Actions matrix pour lancer N jobs en parallèle.
    """
    # Charger toutes les lignes
    all_rows = load_input(input_path, col_orias, col_siren, col_nom)
    total    = len(all_rows)
    size     = -(-total // n_chunks)
    out_dir  = Path(input_path).parent
    stem     = Path(input_path).stem
    ts       = time.strftime("%Y%m%d_%H%M%S")

    chunk_files = []
    for i in range(n_chunks):
        chunk = all_rows[i * size : (i + 1) * size]
        if not chunk:
            break
        fpath = out_dir / f"{stem}_chunk{i+1}.csv"
        with open(fpath, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["orias", "siren", "nom"])
            w.writeheader()
            w.writerows(chunk)
        chunk_files.append(str(fpath))
        print(f"  ✅ Chunk {i+1}/{n_chunks} : {len(chunk)} entrées → {fpath.name}")

    n_actual = len(chunk_files)
    est_h    = (total / n_actual) * 2 * delay / 3600
    print(f"\n  {n_actual} blocs créés | ~{est_h:.1f}h par job (parallèle)")

    # ── Workflow GitHub Actions ──────────────────────────────
    chunk_ids = list(range(1, n_actual + 1))
    workflow  = f"""name: ORIAS Scraper — {n_actual} jobs parallèles

on:
  workflow_dispatch:
    inputs:
      delay:
        description: "Délai DDG en secondes (défaut {delay})"
        default: "{delay}"

jobs:
  scraper:
    runs-on: ubuntu-latest
    timeout-minutes: 360   # 6h max par job

    strategy:
      fail-fast: false      # un job qui échoue n'arrête pas les autres
      matrix:
        chunk: {chunk_ids}

    steps:
      - name: Checkout
        uses: actions/checkout@v4

      - name: Python 3.11
        uses: actions/setup-python@v5
        with:
          python-version: "3.11"
          cache: pip

      - name: Dépendances
        run: |
          pip install aiohttp aiofiles requests beautifulsoup4 lxml \\
                      openpyxl duckduckgo-search tqdm pandas

      - name: Scraping bloc ${{{{ matrix.chunk }}}}/{n_actual}
        env:
          SMTP_PASSWORD: ${{{{ secrets.SMTP_PASSWORD }}}}
        run: |
          python {script_name} run \\
            --input        "{stem}_chunk${{{{ matrix.chunk }}}}.csv" \\
            --output       "results_chunk${{{{ matrix.chunk }}}}_{ts}.xlsx" \\
            --cache        "cache_chunk${{{{ matrix.chunk }}}}.json" \\
            --delay        ${{{{ github.event.inputs.delay }}}} \\
            --email-to     "${{{{ secrets.EMAIL_TO }}}}" \\
            --email-from   "${{{{ secrets.EMAIL_FROM }}}}"

      - name: Upload artefact
        uses: actions/upload-artifact@v4
        if: always()
        with:
          name: results-chunk-${{{{ matrix.chunk }}}}-{ts}
          path: |
            results_chunk${{{{ matrix.chunk }}}}_{ts}.xlsx
            cache_chunk${{{{ matrix.chunk }}}}.json
          retention-days: 30

  merge:
    needs: scraper       # attend que TOUS les jobs soient finis
    runs-on: ubuntu-latest
    if: always()         # fusionne même si certains jobs ont échoué

    steps:
      - name: Checkout
        uses: actions/checkout@v4

      - name: Python
        uses: actions/setup-python@v5
        with:
          python-version: "3.11"
          cache: pip

      - name: Dépendances
        run: pip install openpyxl pandas

      - name: Télécharger tous les artefacts
        uses: actions/download-artifact@v4
        with:
          path: artefacts/

      - name: Fusion des résultats
        run: |
          python {script_name} merge \\
            --pattern "artefacts/**/results_chunk*_{ts}.xlsx" \\
            --output  "orias_resultats_final_{ts}.xlsx"

      - name: Upload résultat final
        uses: actions/upload-artifact@v4
        with:
          name: orias-final-{ts}
          path: orias_resultats_final_{ts}.xlsx
          retention-days: 90

      - name: Envoi email (si configuré)
        if: "${{{{ secrets.EMAIL_TO != '' }}}}"
        env:
          SMTP_PASSWORD: ${{{{ secrets.SMTP_PASSWORD }}}}
        run: |
          python {script_name} notify \\
            --result  "orias_resultats_final_{ts}.xlsx" \\
            --email-to    "${{{{ secrets.EMAIL_TO }}}}" \\
            --email-from  "${{{{ secrets.EMAIL_FROM }}}}"
"""

    wf_dir  = Path(".github/workflows")
    wf_dir.mkdir(parents=True, exist_ok=True)
    wf_path = wf_dir / f"orias_scraper_{ts}.yml"
    wf_path.write_text(workflow, encoding="utf-8")

    print(f"\n  📄 Workflow GitHub Actions : {wf_path}")
    print(f"\n  ÉTAPES :")
    print(f"  1. git add . && git commit -m 'ORIAS scraper {ts}' && git push")
    print(f"  2. GitHub → onglet Actions → 'ORIAS Scraper' → Run workflow")
    print(f"  3. Les {n_actual} jobs tournent en parallèle (~{est_h:.1f}h chacun)")
    print(f"  4. Le job 'merge' fusionne automatiquement à la fin")
    print(f"  5. Téléchargez l'artefact 'orias-final-{ts}' depuis Actions")
    print(f"\n  Durée totale estimée : ~{est_h:.1f}h  "
          f"(vs ~{est_h * n_actual:.1f}h en série)")


# ═══════════════════════════════════════════════════════════
#  MERGE
# ═══════════════════════════════════════════════════════════

def merge_results(pattern: str, output_path: str):
    import glob
    files = sorted(glob.glob(pattern, recursive=True))
    if not files:
        print(f"❌ Aucun fichier trouvé : {pattern}")
        return

    print(f"\n  Fusion de {len(files)} fichier(s) :")
    all_results: list[dict] = []
    seen: set[str] = set()

    for fpath in files:
        # Priorité : fichier cache JSON
        cache_path = str(fpath).replace("results_", "cache_").replace(".xlsx", ".json")
        if Path(cache_path).exists():
            try:
                data = json.loads(Path(cache_path).read_text(encoding="utf-8"))
                for entry in data.values():
                    key = entry.get("numero_orias") or entry.get("siren", "")
                    if key and key not in seen:
                        seen.add(key)
                        all_results.append(entry)
                print(f"  ✅ {Path(fpath).name} → {len(data)} entrées (via cache)")
                continue
            except Exception:
                pass
        # Fallback : lire Excel si pas de cache
        if HAS_PANDAS:
            try:
                df = pd.read_excel(fpath, sheet_name="Résultats ORIAS", skiprows=3, dtype=str)
                n_before = len(all_results)
                for _, row in df.iterrows():
                    key = str(row.iloc[0]).strip()
                    if key and key != "nan" and key not in seen:
                        seen.add(key)
                        all_results.append({
                            "numero_orias": key,
                            "siren": str(row.iloc[1]).strip(),
                            "nom":   str(row.iloc[2]).strip(),
                            "sante_individuelle": "oui" in str(row.iloc[4]).lower(),
                            "perin": "oui" in str(row.iloc[5]).lower(),
                            "score": int(str(row.iloc[6]).strip() or 0),
                            "niveau": str(row.iloc[7]).strip(),
                            "url":    str(row.iloc[8]).strip(),
                            "sources": str(row.iloc[9]).strip(),
                            "methode": "", "detail_score": "", "ape": "", "ville": "",
                        })
                print(f"  ✅ {Path(fpath).name} → {len(all_results)-n_before} entrées (via Excel)")
            except Exception as e:
                print(f"  ⚠  {Path(fpath).name} ignoré : {e}")

    if not all_results:
        print("  ❌ Aucune donnée valide.")
        return

    print(f"\n  Total : {len(all_results)} entrées uniques")
    build_excel(all_results, output_path)


# ═══════════════════════════════════════════════════════════
#  STATS
# ═══════════════════════════════════════════════════════════

def show_stats(cache_path: str):
    cache = load_cache(cache_path)
    if not cache:
        print("Cache vide ou introuvable.")
        return
    n     = len(cache)
    nb_s  = sum(1 for r in cache.values() if r.get("sante_individuelle"))
    nb_p  = sum(1 for r in cache.values() if r.get("perin"))
    nb_u  = sum(1 for r in cache.values() if r.get("url"))
    nb_hi = sum(1 for r in cache.values() if r.get("score", 0) >= 70)
    nb_md = sum(1 for r in cache.values() if 40 <= r.get("score", 0) < 70)
    nb_lo = sum(1 for r in cache.values() if r.get("score", 0) < 40)

    print(f"\n  Cache : {cache_path}")
    print(f"  ──────────────────────────────────")
    print(f"  Entrées traitées    : {n}")
    print(f"  URLs trouvées       : {nb_u}  ({nb_u/n*100:.0f}%)")
    print(f"  Santé Individuelle  : {nb_s}  ({nb_s/n*100:.0f}%)")
    print(f"  PERIN               : {nb_p}  ({nb_p/n*100:.0f}%)")
    print(f"  Score ≥ 70 (fiable) : {nb_hi}")
    print(f"  Score 40–69         : {nb_md}")
    print(f"  Score < 40          : {nb_lo}")


# ═══════════════════════════════════════════════════════════
#  EMAIL NOTIFICATION
# ═══════════════════════════════════════════════════════════

def send_notification(result_path: str, email_to: str, email_from: str,
                      smtp_pwd: str, smtp_host: str, smtp_port: int):
    if not (email_to and email_from and smtp_pwd):
        return
    try:
        cache_files = list(Path(".").glob("cache_chunk*.json"))
        n  = nb_s = nb_p = 0
        for cf in cache_files:
            try:
                data = json.loads(cf.read_text(encoding="utf-8"))
                n    += len(data)
                nb_s += sum(1 for r in data.values() if r.get("sante_individuelle"))
                nb_p += sum(1 for r in data.values() if r.get("perin"))
            except Exception:
                pass

        subject = f"[ORIAS Scraper] Terminé — {nb_s} Santé | {nb_p} PERIN sur {n} courtiers"
        body    = (f"Traitement ORIAS terminé.\n\n"
                   f"  Courtiers analysés : {n}\n"
                   f"  Santé individuelle : {nb_s}\n"
                   f"  PERIN              : {nb_p}\n\n"
                   f"Fichier joint : {Path(result_path).name}")

        msg = EmailMessage()
        msg["Subject"] = subject; msg["From"] = email_from
        msg["To"] = email_to; msg["Date"] = formatdate(localtime=True)
        msg.set_content(body)
        if Path(result_path).exists():
            msg.add_attachment(
                Path(result_path).read_bytes(), maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=Path(result_path).name,
            )
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ssl.create_default_context()) as s:
            s.login(email_from, smtp_pwd); s.send_message(msg)
        print(f"  ✅ Email envoyé à {email_to}")
    except Exception as e:
        print(f"  ❌ Email : {e}")


# ═══════════════════════════════════════════════════════════
#  POINT D'ENTRÉE — SOUS-COMMANDES
# ═══════════════════════════════════════════════════════════

def make_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(
        description="Scraper ORIAS — Santé Individuelle & PERIN",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Commandes :
  run    --input orias.csv [--chunk-id N --total-chunks M]
  split  --input orias.csv --chunks 9
  merge  --pattern "results_chunk_*.xlsx" --output final.xlsx
  stats  --cache cache.json
  notify --result final.xlsx --email-to x@x.com ...
        """,
    )
    sub = ap.add_subparsers(dest="mode")

    def common_args(p):
        p.add_argument("--input",          default="")
        p.add_argument("--col-orias",      default="", help="Colonne numéro ORIAS")
        p.add_argument("--col-siren",      default="", help="Colonne SIREN")
        p.add_argument("--col-nom",        default="", help="Colonne raison sociale")
        return p

    # run
    p_run = common_args(sub.add_parser("run"))
    p_run.add_argument("--output",       default="")
    p_run.add_argument("--cache",        default="orias_cache.json")
    p_run.add_argument("--chunk-id",     default=0, type=int,   help="N° du bloc (matrix GitHub)")
    p_run.add_argument("--total-chunks", default=1, type=int,   help="Nb total de blocs")
    p_run.add_argument("--workers",      default=3, type=int,   help="Threads DDG (défaut 3)")
    p_run.add_argument("--delay",        default=1.5, type=float)
    p_run.add_argument("--reset",        action="store_true")
    p_run.add_argument("--email-to",     default="")
    p_run.add_argument("--email-from",   default="")
    p_run.add_argument("--smtp-password",default="")
    p_run.add_argument("--smtp-host",    default="smtp.gmail.com")
    p_run.add_argument("--smtp-port",    default=465, type=int)

    # split
    p_spl = common_args(sub.add_parser("split"))
    p_spl.add_argument("--chunks",  default=9, type=int)
    p_spl.add_argument("--delay",   default=1.5, type=float)

    # merge
    p_mrg = sub.add_parser("merge")
    p_mrg.add_argument("--pattern", default="results_chunk_*.xlsx")
    p_mrg.add_argument("--output",  default="orias_resultats_final.xlsx")

    # stats
    p_sta = sub.add_parser("stats")
    p_sta.add_argument("--cache", default="orias_cache.json")

    # notify
    p_ntf = sub.add_parser("notify")
    p_ntf.add_argument("--result",        default="")
    p_ntf.add_argument("--email-to",      default="")
    p_ntf.add_argument("--email-from",    default="")
    p_ntf.add_argument("--smtp-password", default="")
    p_ntf.add_argument("--smtp-host",     default="smtp.gmail.com")
    p_ntf.add_argument("--smtp-port",     default=465, type=int)

    return ap


if __name__ == "__main__":
    ap   = make_parser()
    args = ap.parse_args()

    if args.mode == "split":
        if not args.input:
            print("❌ --input requis"); exit(1)
        print("=" * 60)
        print(f"   Split + Workflow GitHub Actions ({args.chunks} jobs)")
        print("=" * 60)
        split_and_generate_workflow(
            args.input, args.chunks,
            col_orias=args.col_orias, col_siren=args.col_siren, col_nom=args.col_nom,
            delay=args.delay,
        )

    elif args.mode == "merge":
        print("=" * 60); print("   Merge"); print("=" * 60)
        merge_results(args.pattern, args.output)

    elif args.mode == "stats":
        show_stats(args.cache)

    elif args.mode == "notify":
        smtp_pwd = args.smtp_password or os.environ.get("SMTP_PASSWORD", "")
        send_notification(args.result, args.email_to, args.email_from,
                          smtp_pwd, args.smtp_host, args.smtp_port)

    else:   # run (défaut)
        if not args.input:
            print("❌ --input requis"); exit(1)

        ts         = time.strftime("%Y%m%d_%H%M%S")
        chunk_sfx  = f"_chunk{args.chunk_id}" if args.chunk_id else ""
        out_path   = args.output or f"results{chunk_sfx}_{ts}.xlsx"
        cache_path = args.cache  or f"cache{chunk_sfx}.json"

        print("=" * 60)
        print(f"   ORIAS Scraper — run{' (bloc '+str(args.chunk_id)+'/'+str(args.total_chunks)+')' if args.chunk_id else ''}")
        print("=" * 60)
        print(f"  Entrée  : {args.input}")
        print(f"  Sortie  : {out_path}")
        print(f"  Cache   : {cache_path}")
        print(f"  Threads : {args.workers}  |  Délai DDG : {args.delay}s")
        if not HAS_AIOHTTP:
            print("  ⚠  aiohttp non disponible — mode synchrone (plus lent)")

        if args.reset and Path(cache_path).exists():
            Path(cache_path).unlink()

        rows  = load_input(args.input, args.col_orias, args.col_siren, args.col_nom,
                           args.chunk_id, args.total_chunks)
        cache = load_cache(cache_path)

        to_do = sum(1 for r in rows if r["orias"] not in cache)
        est_h = to_do * 2 * args.delay / 3600
        print(f"\n  {to_do} à traiter | durée estimée : ~{est_h:.1f}h\n")

        results = asyncio.run(
            process_chunk(rows, cache, cache_path, args.delay, args.workers)
        )

        print("\nGénération Excel…")
        build_excel(results, out_path, args.chunk_id, args.total_chunks)

        smtp_pwd = args.smtp_password or os.environ.get("SMTP_PASSWORD", "")
        if args.email_to and args.email_from and smtp_pwd:
            send_notification(out_path, args.email_to, args.email_from,
                              smtp_pwd, args.smtp_host, args.smtp_port)

        print("\nTerminé ✅")